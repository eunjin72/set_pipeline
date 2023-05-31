import os
import re

from openpyxl import *
from openpyxl.drawing.image import Image
import OpenEXR

from pprint import pprint
from ffmpeg import *


class ExcelCreater:

    def __init__(self):

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Shot'

        self._input_path = None
        self._output_path = None

        self.files_dict = {}

        self.first_file_list = []
        self.last_file_list = []

        # self.start_meta = None
        # self.last_meta = None

        self.exr_meta_list = []

        self.img_file_list = []

        self.project_path = self.input_path.split('/production/scan')

    @property
    def input_path(self):
        return self._input_path

    @input_path.setter
    def input_path(self, value):
        if value is None or value == "":
            raise ValueError("Input path is missing.")
        self._input_path = value

    # @property
    # def output_path(self):
    #     return self._output_path

    # @output_path.setter
    # def output_path(self, value):
    #     self._output_path = value

    def get_all_files(self):
        self.files_dict = {}

        for root, dirs, files in os.walk(self.input_path):
            if files:
                # print(f"ffff==={files}")
                files.sort(key=lambda x: int(re.findall(r'\d+', x)[-1]))
                self.files_dict[root] = files
        if len(self.files_dict.values()) == 0:
            raise Exception("No files found in the directory.")
        # pprint(f"olol==={self.files_dict}")

        return self.files_dict

    def get_first_and_last_file(self):
        self.first_file_list = []
        self.last_file_list = []
        for root, files in self.files_dict.items():
            if len(files) > 0:
                self.first_file_list.append(root + "/" + files[0])
                self.last_file_list.append(root + "/" + files[-1])
        # pprint(f"111=={self.first_file_list}, 222=={self.last_file_list}")
        return self.first_file_list, self.last_file_list

    def get_meta(self):
        # self.origin_data()
        for i, exr in enumerate(self.first_file_list):
            ######test######
            # exr.sort(reverse=False)
            # pprint(f"bb=={i}=={exr}")

            exr_start_file = OpenEXR.InputFile(exr)
            # self.start_meta = exr_start_file.header()
            start_meta = exr_start_file.header()

            exr_last_file = OpenEXR.InputFile(exr)
            # self.last_meta = exr_last_file.header()
            last_meta = exr_last_file.header()
            # print(f"333=={self.start_meta}, 444=={self.last_meta}")

            file_data = re.match(r"(.*/)([^/]+)\.(\d+)\.(\w+)$", exr)

            # 해상도
            # res = re.findall(r'\d+\d+', str(self.start_meta.get("dataWindow")))
            res = re.findall(r'\d+\d+', str(start_meta.get("dataWindow")))
            resolutions = list(map(lambda x: str(int(x) + 1), res))

            # 프레임
            # frames = re.findall(r'\d+\.\d+|\d+', str(self.start_meta.get("framesPerSecond")))
            frames = re.findall(r'\d+\.\d+|\d+', str(start_meta.get("framesPerSecond")))

            self.exr_meta_list.append(
                {
                    "scan_path": file_data.group(1),
                    "scan_name": file_data.group(2),
                    # "clip_name": self.start_meta.get("interim.clip.cameraClipName"),
                    "clip_name": start_meta.get("interim.clip.cameraClipName"),
                    "pad": '%0' + str(len(file_data.group(3))) + 'd',
                    "ext": file_data.group(4),
                    "resolutions": ' x '.join(resolutions),
                    "start_frame": int(frames[1]),
                    "and_frame": int(frames[0]),
                    "duration": int(frames[0]) - int(frames[1]) + 1,
                    # "timecode_in": self.start_meta.get("arriraw/timeCode"),
                    "timecode_in": start_meta.get("arriraw/timeCode"),
                    # "timecode_out": self.last_meta.get("arriraw/timeCode"),
                    "timecode_out": last_meta.get("arriraw/timeCode"),
                    "framerate": float(frames[2]),
                    # "date": self.start_meta.get("capDate")
                    "date": start_meta.get("capDate")
                }
            )
        # pprint(f"wvwv===={self.exr_meta_list}")

    def get_thumbnail(self):
        self.thumbnail_path = os.path.join(self.project_path[0], f'tmp/thumb{self.project_path[1]}')
        if not os.path.exists(self.thumbnail_path):
            os.makedirs(self.thumbnail_path, exist_ok=True)

        exr_files_dict = {}
        for path, dirs, files in os.walk(self.input_path):
            if len(files) > 0:
                files.sort(reverse=False)
                names = files[0].split('.exr')[0]
                exr_files_dict[os.path.join(path, files[0])] = names

        # convert exr to jpg
        for exr_file, file_name in exr_files_dict.items():
            run(output(input(exr_file), f'{self.thumbnail_path}/{file_name}.jpg'))

    def insert_thumbnail(self):
        thumbnail_lists = os.listdir(self.thumbnail_path)
        for i, thumbnail_list in enumerate(thumbnail_lists):
            image = Image(os.path.join(self.thumbnail_path, thumbnail_list))
            image.width = 250
            image.height = 150
            col_width = image.width * 50 / 350   ## 엑셀 셀 폭 높이 단위
            row_height = image.height * 250 / 300
            self.ws.add_image(image, anchor='B' + str(i + 2))  ## 이미지 삽입
            if i == 0:
                self.ws.column_dimensions['B'].width = col_width  ## 셀 폭은 한 번만 변경
            self.ws.row_dimensions[i + 2].height = row_height  ## 셀 높이 변경
            self.ws.cell(row=i + 2, column=2, value=thumbnail_list) ## file_name 입력

    def execl_form(self):
        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]
        for i, title in enumerate(header_list):
            # print(f"titi=={i}")
            self.ws.cell(row=1, column=i + 1, value=title)

    def excel_create(self):

        self.execl_form()
        self.get_thumbnail()
        self.insert_thumbnail()
        self.get_meta()

        # for row, meta in enumerate(self.exr_meta_list, start=2):
            # print(f"coco=={c}==={meta}")

            # self.ws.cell(row=row, column=8, value=meta.get("scan_path"))
            # self.ws.cell(row=row, column=9, value=meta.get("scan_name"))
            # self.ws.cell(row=row, column=10, value=meta.get("clip_name"))
            # self.ws.cell(row=row, column=11, value=meta.get("pad"))
            # self.ws.cell(row=row, column=12, value=meta.get("ext"))
            # self.ws.cell(row=row, column=13, value=meta.get("resolutions"))
            # self.ws.cell(row=row, column=14, value=meta.get("start_frame"))
            # self.ws.cell(row=row, column=15, value=meta.get("and_frame"))
            # self.ws.cell(row=row, column=16, value=meta.get("duration"))

            # self.ws.cell(row=row, column=20, value=meta.get("timecode_in"))
            # self.ws.cell(row=row, column=21, value=meta.get("timecode_out"))

            # self.ws.cell(row=row, column=24, value=meta.get("framerate"))
            # self.ws.cell(row=row, column=25, value=meta.get("date"))

        #### test #####
        for meta in self.exr_meta_list:
            self.ws.append(  
            {
                'H': meta.get("scan_path"),
                'I': meta.get("scan_name"),
                'J': meta.get("clip_name"),
                'K': meta.get("pad"),
                'L': meta.get("ext"),
                'M': meta.get("resolutions"),
                'N': meta.get("start_frame"),
                'O': meta.get("and_frame"),
                'P': meta.get("duration"),
                'T': meta.get("timecode_in"),
                'U': meta.get("timecode_out"),
                'X': meta.get("framerate"),
                'Y': meta.get("date")
                })

        self.excel_save()

    def excel_save(self):
        excel_path = os.path.join(self.project_path[0], 'production/excel')
        # name = self.project_path[1] + '.csv'
        name = self.project_path[1] + '.xlsx'
        save_dir_path = os.path.join(excel_path, name)

        count = 1
        while os.path.exists(save_dir_path):
            # new_name = f'{self.project_path[1]}_{count}.csv'
            new_name = f'{self.project_path[1]}_{count}.xlsx'
            save_dir_path = os.path.join(excel_path, new_name)
            count += 1
        
        self.wb.save(save_dir_path)


def main():
    ec = ExcelCreater()

    # setter test info
    ec.input_path = r"/TD/show/hanjin/production/scan/20221017_plate_scan"

    ec.get_all_files()
    ec.get_first_and_last_file()

    print(f"mack{ec.excel_create()}")


if __name__ == '__main__':
    main()
